import tkinter as tk
from tkinter import ttk, Toplevel, Frame, messagebox, Listbox, MULTIPLE
from tkcalendar import Calendar
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Border, Side
import sqlite3


def initialize_db():
    with sqlite3.connect('employees.db') as conn:
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS employees (name TEXT UNIQUE)")
        cursor.execute("CREATE TABLE IF NOT EXISTS schedules (date TEXT, shift TEXT, employee TEXT)")
        conn.commit()

initialize_db()


def update_employee_selection(current_date, excluding_listbox=None):
    all_selected = []
    for shift, lb in shift_employees[current_date].items():
        if lb != excluding_listbox:
            all_selected.extend(lb.get(0, tk.END))
    available_employees = [emp for emp in all_workers if emp not in all_selected]
    return available_employees

def add_to_shift(shift_listbox, current_date, shift_time):
    def add_selections():
        selected_employees = [employee_list.get(i) for i in employee_list.curselection()]
        for emp in selected_employees:
            with sqlite3.connect('employees.db') as conn:
                cursor = conn.cursor()
                cursor.execute("INSERT INTO schedules (date, shift, employee) VALUES (?, ?, ?)", (current_date, shift_time, emp))
                conn.commit()
            shift_listbox.insert(tk.END, emp)
        selection_window.destroy()

    available_employees = update_employee_selection(current_date, excluding_listbox=shift_listbox)

    # Subtract the currently selected employees from the available list
    current_selected = shift_listbox.get(0, tk.END)
    available_employees = [emp for emp in available_employees if emp not in current_selected]

    selection_window = Toplevel(root)
    employee_list = Listbox(selection_window, selectmode=MULTIPLE, exportselection=False)
    employee_list.pack(fill=tk.BOTH, expand=True)
    for worker in available_employees:
        employee_list.insert(tk.END, worker)
    btn_add = ttk.Button(selection_window, text="Add Selected", command=add_selections)
    btn_add.pack()


def manage_employees():
    def add_from_entry():
        new_employee = new_employee_entry.get()
        with sqlite3.connect('employees.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM employees WHERE name=?", (new_employee,))
            existing_employee = cursor.fetchone()

            if new_employee and not existing_employee:
                cursor.execute("INSERT INTO employees (name) VALUES (?)", (new_employee,))
                conn.commit()
                employees_listbox.insert(tk.END, new_employee)
                new_employee_entry.delete(0, tk.END)
            elif existing_employee:
                messagebox.showwarning("Warning", f"{new_employee} already exists!")
            else:
                messagebox.showwarning("Warning", "Employee name cannot be empty!")

    def remove_selected():
        selected_indices = list(employees_listbox.curselection())
        with sqlite3.connect('employees.db') as conn:
            cursor = conn.cursor()
            for index in reversed(selected_indices):
                employee_to_remove = employees_listbox.get(index)
                cursor.execute("DELETE FROM employees WHERE name=?", (employee_to_remove,))
                employees_listbox.delete(index)
            conn.commit()

    manage_window = Toplevel(root)
    manage_window.title("Manage Employees")
    employees_listbox = Listbox(manage_window, selectmode=MULTIPLE)

    with sqlite3.connect('employees.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM employees")
        all_workers_db = cursor.fetchall()
        for worker in all_workers_db:
            employees_listbox.insert(tk.END, worker[0])

    employees_listbox.pack(padx=10, pady=10)
    new_employee_entry = ttk.Entry(manage_window)
    new_employee_entry.pack(side=tk.LEFT, padx=5)
    add_button = ttk.Button(manage_window, text="+", command=add_from_entry)
    add_button.pack(side=tk.LEFT, padx=5)
    remove_button = ttk.Button(manage_window, text="-", command=remove_selected)
    remove_button.pack(side=tk.LEFT, padx=5)

def remove_from_shift(shift_listbox, current_date, shift_time):
    selected_indices = list(shift_listbox.curselection())
    for index in reversed(selected_indices):
        with sqlite3.connect('employees.db') as conn:
            cursor = conn.cursor()
            employee_to_remove = shift_listbox.get(index)
            cursor.execute("DELETE FROM schedules WHERE date=? AND shift=? AND employee=?", (current_date, shift_time, employee_to_remove))
            conn.commit()
        shift_listbox.delete(index)


def show_schedule(start_date, end_date):
    schedule_window = Toplevel(root)

    with sqlite3.connect('employees.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM employees")
        all_workers_db = cursor.fetchall()
        global all_workers
        all_workers = [worker[0] for worker in all_workers_db]

    header_frame = ttk.Frame(schedule_window)
    header_frame.pack(fill=tk.X)
    btn_export_schedule = ttk.Button(header_frame, text="Export to Excel", command=export_to_excel)
    btn_export_schedule.pack(pady=10)

    canvas = tk.Canvas(schedule_window)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar = ttk.Scrollbar(schedule_window, orient="vertical", command=canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    canvas.configure(yscrollcommand=scrollbar.set)

    frame_content = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=frame_content, anchor='nw')

    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    global shift_employees
    shift_employees = {}

    shift_colors = {
        "12-9": "#BDBAB3",
        "16-1": "#F8D7C1",
        "8-17": "#D9F5F8",
        "Leave": "#F8B4B4",
        "OFF-Day": "#E4CC98",
    }

    for idx, current_date in enumerate(date_range):
        if current_date.weekday() == 4:  # If the day is a Friday
            frame = tk.LabelFrame(frame_content, text=current_date.strftime('%Y-%m-%d'), padx=5, pady=5, bg="#E6EDB8")

        else:
            frame = tk.LabelFrame(frame_content, text=current_date.strftime('%Y-%m-%d'), padx=5, pady=5)

        frame.grid(row=idx // 9, column=idx % 9)
        shift_employees[current_date] = {}

        for shift in ["12-9", "16-1", "8-17", "OFF-Day", "Leave"]:
            lbl_shift = tk.Label(frame, text=shift)
            lbl_shift.pack()

            bg_color = shift_colors.get(shift, "white")
            shift_listbox = Listbox(frame, height=4, selectmode=MULTIPLE, bg=bg_color)
            shift_listbox.pack()

            with sqlite3.connect('employees.db') as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT employee FROM schedules WHERE date=? AND shift=?",
                               (current_date.strftime('%Y-%m-%d'), shift))
                employees_for_this_shift = cursor.fetchall()
                for employee_tuple in employees_for_this_shift:
                    shift_listbox.insert(tk.END, employee_tuple[0])

            shift_employees[current_date][shift] = shift_listbox

            control_frame = Frame(frame)  # Container frame for the + and - buttons
            control_frame.pack()
            btn_add_employee = ttk.Button(control_frame, text="+",
                                          command=lambda lb=shift_listbox, dt=current_date, st=shift: add_to_shift(lb,
                                                                                                                   dt,
                                                                                                                   st))
            btn_add_employee.pack(side=tk.LEFT)
            btn_remove_employee = ttk.Button(control_frame, text="-", command=lambda lb=shift_listbox, dt=current_date,
                                                                                     st=shift: remove_from_shift(lb, dt,
                                                                                                                 st))
            btn_remove_employee.pack(side=tk.LEFT)

    frame_content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))


def refresh_listboxes(current_date):
    for shift, lb in shift_employees[current_date].items():
        current_selected = lb.get(0, tk.END)
        available_for_this_shift = update_employee_selection(current_date, excluding_listbox=lb)
        # Now, we need to make sure we don't add employees that are already in this shift listbox
        for worker in available_for_this_shift:
            if worker not in current_selected:
                lb.insert(tk.END, worker)


def export_to_excel():
    # Create a new Excel workbook and add a worksheet to it
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Style for the headers
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Headers
    ws["A1"] = "Date"
    ws["A1"].font = bold_font
    ws["B1"] = "Shift"
    ws["B1"].font = bold_font
    ws["C1"] = "Employees"
    ws["C1"].font = bold_font

    row = 2

    # Iterate through the dates and shifts, and write them to the Excel sheet
    for date, shifts in shift_employees.items():
        for shift, employees in shifts.items():
            ws.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=1).border = thin_border  # Set border for date cell

            ws.cell(row=row, column=2, value=shift)
            ws.cell(row=row, column=2).border = thin_border  # Set border for shift cell

            ws.cell(row=row, column=3, value=", ".join(employees.get(0, tk.END)))
            ws.cell(row=row, column=3).border = thin_border  # Set border for employees cell

            row += 1

    # Adjust column widths for a better appearance
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[cell.column_letter].width = adjusted_width

    # Save the workbook
    filename = "schedule.xlsx"
    wb.save(filename)
    messagebox.showinfo("Info", f"Schedule exported to {filename}")

def on_date_selection():
    show_schedule(cal_start.selection_get(), cal_end.selection_get())

root = tk.Tk()
root.title("Schedule Selector")
all_workers = ["", "", "", ""]
btn_manage_employees = ttk.Button(root, text="Manage Employees", command=manage_employees)
btn_manage_employees.pack(pady=(20, 10))
cal_start = Calendar(root, selectmode="day", date_pattern='y-mm-dd', background='darkblue', foreground='white', tooltipforeground='black')
cal_start.pack(padx=10, pady=10)
cal_end = Calendar(root, selectmode="day", date_pattern='y-mm-dd', background='darkblue', foreground='white', tooltipforeground='black')
cal_end.pack(padx=10, pady=10)
btn_submit = ttk.Button(root, text="Show Schedule", command=on_date_selection)
btn_submit.pack(pady=20)
root.mainloop()
